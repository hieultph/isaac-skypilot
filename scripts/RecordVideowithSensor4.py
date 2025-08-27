# Lưu video từ camera Airsim

import airsim # Thư viện client của AirSim
import numpy as np # Dành cho các phép toán số, đặc biệt là xử lý hình ảnh
import cv2 # Thư viện OpenCV để ghi video và xử lý hình ảnh
import time # Dành cho các hoạt động dựa trên thời gian như độ trễ và đo lường hiệu suất
import os # Dành cho các tương tác với hệ điều hành

# --- Biến Toàn cục ---
AIRSIM_CLIENT = None
# Tên camera mặc định. Thay đổi nếu bạn sử dụng các camera có tên cụ thể trong cài đặt AirSim của mình.
CAMERA_NAME = "0"
# Loại hình ảnh cần chụp.
IMAGE_TYPE = airsim.ImageType.Scene # Hình ảnh màu tiêu chuẩn

# --- Cài đặt Ghi Video ---
# OUTPUT_FILENAME hiện được tạo động trong khối __main__ để đảm bảo tính duy nhất
PLAYBACK_FPS = 30.0 # FPS mong muốn cho việc PHÁT LẠI tệp video đầu ra.
                    # Điều này quyết định tốc độ khung hình và nhịp độ thời gian thực của video đầu ra.
FOURCC = cv2.VideoWriter_fourcc(*'mp4v') # Dành cho tệp .mp4. 'X264' là một lựa chọn thay thế tuyệt vời nếu có sẵn.
VIDEO_WRITER = None

# --- Biến Tính toán FPS (để theo dõi tốc độ chụp thực tế) ---
capture_fps_start_time = time.time()
capture_fps_frame_count = 0
current_capture_fps = 0

def connect_to_airsim():
    """Kết nối với client của AirSim."""
    global AIRSIM_CLIENT
    try:
        AIRSIM_CLIENT = airsim.MultirotorClient()
        # AIRSIM_CLIENT = airsim.CarClient() # Dành cho mô phỏng xe hơi
        AIRSIM_CLIENT.confirmConnection()
        print("Đã kết nối thành công với AirSim.")
    except Exception as e:
        print(f"Lỗi khi kết nối với AirSim: {e}")
        print("Vui lòng đảm bảo mô phỏng AirSim đang chạy và có thể truy cập được.")
        AIRSIM_CLIENT = None

# --- Thiết lập xử lý lỗi AirSim RpcError mạnh mẽ ---
try:
    from airsim.types import RpcError as AirSimRpcError
except ImportError:
    AirSimRpcError = Exception # Sử dụng Exception chung để tương thích

def get_image_type_name(img_type_enum_or_int):
    """Lấy tên chuỗi của một ImageType của AirSim một cách an toàn."""
    if hasattr(img_type_enum_or_int, 'name'):
        return img_type_enum_or_int.name
    elif isinstance(img_type_enum_or_int, int):
        type_map = {0: "Scene", 1: "DepthPerspective", 2: "DepthPlanar", 5: "Segmentation"}
        return type_map.get(img_type_enum_or_int, str(img_type_enum_or_int))
    return str(img_type_enum_or_int)

def record_airsim_video():
    """
    Ghi video từ AirSim vào một tệp cho đến khi nhấn Ctrl+C, đảm bảo tốc độ thời gian thực.
    """
    global AIRSIM_CLIENT, VIDEO_WRITER
    global capture_fps_start_time, capture_fps_frame_count, current_capture_fps

    if not AIRSIM_CLIENT:
        print("Client AirSim không khả dụng. Không thể bắt đầu ghi.")
        return

    print(f"Đang cố gắng ghi video vào {globals().get('OUTPUT_FILENAME', 'airsim_video.mp4')} (FPS phát lại: {PLAYBACK_FPS}).")
    print("Nhấn Ctrl+C để dừng ghi.")
    
    frames_written = 0
    last_image_dims = None
    last_recorded_frame = None # Để giữ khung hình cuối cùng được chụp thành công từ AirSim

    # --- Biến đồng bộ hóa thời gian thực ---
    start_recording_time = time.time()
    next_video_frame_write_time = start_recording_time
    desired_video_frame_interval = 1.0 / PLAYBACK_FPS

    capture_fps_start_time = time.time() # Đặt lại để theo dõi FPS trên console

    # --- Prepare sensor data logging ---
    sensor_log_filename = globals().get('OUTPUT_FILENAME', 'airsim_video.mp4').replace('.mp4', '_sensor_log.xlsx')
    import openpyxl
    from openpyxl.styles import Alignment, Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AirSim Sensor Log'
    
    # === SENSOR CONFIGURATION REFERENCE ===
    # Add sensor names used in this script for settings.json reference
    ws.append(['SENSOR NAMES USED (for settings.json reference):'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws['A1'].font = Font(bold=True, color='FF0000')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    sensor_names_info = [
        'State: getMultirotorState()', 'RC: getRCData()', 'Kinematics: simGetGroundTruthKinematics()',
        'Camera: simGetCameraPose()', 'Required for: Position, Velocity, Gimbal, Drone orientation'
    ]
    ws.append(sensor_names_info)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
    ws['A2'].font = Font(italic=True, color='FF666666')
    
    # Add a comment row for context
    ws.append(['Simplified AirSim Log: Frame, Time, Position(X,Y,Z), Velocity(X,Y,Z), Gimbal(R,P,Y), DroneOri(R,P,Y,T)'])
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=19)
    ws['A3'].font = Font(bold=True)
    ws['A3'].alignment = Alignment(horizontal='center')
    # Simplified sensor headers (2 rows)
    group_row = [
        'Frame', 'Frame', 'Frame',
        'Position (NED)', 'Position (NED)', 'Position (NED)',
        'Velocity (NED)', 'Velocity (NED)', 'Velocity (NED)',
        'Gimbal (Camera)', 'Gimbal (Camera)', 'Gimbal (Camera)',
        'Drone Orientation', 'Drone Orientation', 'Drone Orientation', 'Drone Orientation'
    ]
    ws.append(group_row)
    header_row = [
        'Frame', 'Video Time (s)', 'System Time (s)',
        'X (m)', 'Y (m)', 'Z (m)',
        'X (m/s)', 'Y (m/s)', 'Z (m/s)',
        'Gimbal Roll (deg)', 'Gimbal Pitch (deg)', 'Gimbal Yaw (deg)',
        'Drone Roll (deg)', 'Drone Pitch (deg)', 'Drone Yaw (deg)', 'Drone Throttle (deg)'
    ]
    ws.append(header_row)
    # Freeze panes below header
    ws.freeze_panes = 'A6'
    try:
        # --- Thu nhận Khung hình Ban đầu (Bắt buộc để lấy kích thước video) ---
        print("Đang thu nhận khung hình ban đầu để thiết lập trình ghi video...")
        initial_frame_acquired = False
        while not initial_frame_acquired:
            try:
                request = airsim.ImageRequest(CAMERA_NAME, IMAGE_TYPE, False, False)
                responses = AIRSIM_CLIENT.simGetImages([request])

                if responses and responses[0].image_data_uint8:
                    response = responses[0]
                    actual_height = response.height
                    actual_width = response.width

                    if actual_height > 0 and actual_width > 0:
                        img_np = np.frombuffer(response.image_data_uint8, dtype=np.uint8)
                        img_bgr = None

                        # --- ĐÃ SỬA ĐỔI: Logic Xử lý Hình ảnh từ Mã Truyền phát ---
                        if IMAGE_TYPE == airsim.ImageType.Scene:
                            expected_size_bgr = actual_height * actual_width * 3

                            if len(img_np) == expected_size_bgr:
                                # Dữ liệu là 3 kênh (ví dụ: BGR). Định hình lại trực tiếp.
                                img_bgr = img_np.reshape(actual_height, actual_width, 3)
                            else:
                                print(f"Cảnh báo: Kích thước dữ liệu khung hình ban đầu không khớp. Dự kiến {expected_size_bgr}, nhận được {len(img_np)}. Đang thử lại...")
                        # --- KẾT THÚC SỬA ĐỔI ---
                        
                        if img_bgr is not None:
                            last_recorded_frame = img_bgr
                            last_image_dims = (actual_width, actual_height)
                            initial_frame_acquired = True
                            print(f"Đã thu nhận khung hình ban đầu. Kích thước: {last_image_dims}")
                        else:
                            print("Cảnh báo: Không xử lý được hình ảnh ban đầu. Đang thử lại...")
                            time.sleep(0.1)
                    else:
                        print(f"Cảnh báo: Kích thước không hợp lệ cho hình ảnh ban đầu. Đang thử lại...")
                        time.sleep(0.1)
                else:
                    print("Cảnh báo: Không nhận được dữ liệu hình ảnh ban đầu. Đang thử lại...")
                    time.sleep(0.5)
            except AirSimRpcError as rpc_error:
                print(f"Lỗi RPC ban đầu: {rpc_error}. Đang thử lại...")
                time.sleep(1.0)
            except Exception as e:
                print(f"Lỗi không mong muốn trong quá trình thu nhận khung hình ban đầu: {e}. Đang thử lại...")
                time.sleep(0.5)
        
        if last_recorded_frame is None or last_image_dims is None:
            print("Lỗi nghiêm trọng: Không thể thu nhận khung hình ban đầu. Đang thoát.")
            return

        # Khởi tạo VIDEO_WRITER sau khi thu nhận khung hình ban đầu
        VIDEO_WRITER = cv2.VideoWriter(globals()['OUTPUT_FILENAME'], FOURCC, PLAYBACK_FPS, last_image_dims)
        if not VIDEO_WRITER.isOpened():
            print(f"Lỗi: Không thể mở trình ghi video cho {globals()['OUTPUT_FILENAME']}")
            return

        print(f"\rTrình ghi video đã được khởi tạo. Đang ghi vào {globals()['OUTPUT_FILENAME']}")

        # --- Vòng lặp Ghi chính ---
        while True:
            current_real_time = time.time()
            frames_to_write_this_iter = 0
            while next_video_frame_write_time <= current_real_time:
                frames_to_write_this_iter += 1
                next_video_frame_write_time += desired_video_frame_interval
            
            new_frame_successfully_acquired = False
            try:
                request = airsim.ImageRequest(CAMERA_NAME, IMAGE_TYPE, False, False)
                responses = AIRSIM_CLIENT.simGetImages([request])

                if responses and responses[0].image_data_uint8:
                    response = responses[0]
                    actual_height = response.height
                    actual_width = response.width

                    if actual_height > 0 and actual_width > 0:
                        img_np_current = np.frombuffer(response.image_data_uint8, dtype=np.uint8)
                        current_img_bgr = None
                        
                        # --- ĐÃ SỬA ĐỔI: Logic Xử lý Hình ảnh từ Mã Truyền phát ---
                        if IMAGE_TYPE == airsim.ImageType.Scene:
                            expected_size_bgr = actual_height * actual_width * 3
                            current_img_bgr = img_np_current.reshape(actual_height, actual_width, 3)


                        if current_img_bgr is not None:
                            last_recorded_frame = current_img_bgr
                            new_frame_successfully_acquired = True
                            # Xử lý thay đổi kích thước giữa quá trình ghi
                            if last_image_dims != (actual_width, actual_height):
                                print(f"\nKích thước hình ảnh đã thay đổi. Đang khởi tạo lại trình ghi.")
                                last_image_dims = (actual_width, actual_height)
                                if VIDEO_WRITER.isOpened(): VIDEO_WRITER.release()
                                VIDEO_WRITER = cv2.VideoWriter(globals()['OUTPUT_FILENAME'], FOURCC, PLAYBACK_FPS, last_image_dims)
                                if not VIDEO_WRITER.isOpened():
                                    print(f"\nLỗi: Không thể mở lại trình ghi video.")
                                    break
                        else:
                            print("Cảnh báo: Không xử lý được hình ảnh hiện tại sang BGR. Sử dụng lại khung hình cuối.")
                else:
                    print("Cảnh báo: Không nhận được hình ảnh hợp lệ. Sử dụng lại khung hình cuối.")
            
            except AirSimRpcError as rpc_error:
                print(f"\nLỗi AirSim RPC: {rpc_error}. Đang thoát khỏi quá trình ghi.")
                break
            except Exception as e:
                print(f"\nLỗi không mong muốn trong quá trình lấy khung hình: {e}. Sử dụng lại khung hình cuối.")

            # --- Ghi Khung hình vào Trình ghi Video và log sensor data ---
            if last_recorded_frame is not None:
                if not VIDEO_WRITER.isOpened():
                    print("Trình ghi video chưa mở. Đang dừng.")
                    break
                for _ in range(frames_to_write_this_iter):
                    VIDEO_WRITER.write(last_recorded_frame)
                    frames_written += 1
                    # --- Log simplified sensor data for each video frame ---
                    try:
                        # Get essential sensor data
                        try:
                            state = AIRSIM_CLIENT.getMultirotorState(vehicle_name="PX4")
                        except Exception:
                            state = None
                        
                        # RC data for control inputs
                        try:
                            rc = AIRSIM_CLIENT.getRCData(vehicle_name="PX4")
                        except Exception:
                            rc = None

                        now = time.time()
                        video_time = frames_written / PLAYBACK_FPS
                        
                        # Position/velocity from ground truth kinematics
                        try:
                            kinematics = AIRSIM_CLIENT.simGetGroundTruthKinematics(vehicle_name="PX4")
                            pos = kinematics.position
                            vel = kinematics.linear_velocity
                            drone_ori = kinematics.orientation  # This is drone body orientation
                        except Exception as e:
                            # Fallback to state kinematics if ground truth fails
                            pos = state.kinematics_estimated.position if state else None
                            vel = state.kinematics_estimated.linear_velocity if state else None
                            drone_ori = state.kinematics_estimated.orientation if state else None

                        # Get camera gimbal pose (separate from drone orientation)
                        gimbal_roll, gimbal_pitch, gimbal_yaw = 0, 0, 0
                        try:
                            camera_pose = AIRSIM_CLIENT.simGetCameraPose(vehicle_name="PX4", camera_name=CAMERA_NAME)
                            if camera_pose and hasattr(camera_pose, 'orientation'):
                                # Convert camera quaternion to Euler angles
                                import math
                                w = camera_pose.orientation.w_val
                                x = camera_pose.orientation.x_val
                                y = camera_pose.orientation.y_val
                                z = camera_pose.orientation.z_val
                                
                                # Roll (x-axis rotation)
                                sinr_cosp = 2 * (w * x + y * z)
                                cosr_cosp = 1 - 2 * (x * x + y * y)
                                gimbal_roll = math.degrees(math.atan2(sinr_cosp, cosr_cosp))
                                
                                # Pitch (y-axis rotation)
                                sinp = 2 * (w * y - z * x)
                                if abs(sinp) >= 1:
                                    gimbal_pitch = math.degrees(math.copysign(math.pi / 2, sinp))
                                else:
                                    gimbal_pitch = math.degrees(math.asin(sinp))
                                
                                # Yaw (z-axis rotation)
                                siny_cosp = 2 * (w * z + x * y)
                                cosy_cosp = 1 - 2 * (y * y + z * z)
                                gimbal_yaw = math.degrees(math.atan2(siny_cosp, cosy_cosp))
                        except Exception as e:
                            # If camera pose fails, use zeros
                            pass

                        # Convert drone quaternion to Euler angles (drone body orientation)
                        import math
                        drone_roll, drone_pitch, drone_yaw = 0, 0, 0
                        if drone_ori:
                            # Convert drone quaternion to Euler angles in degrees
                            w, x, y, z = drone_ori.w_val, drone_ori.x_val, drone_ori.y_val, drone_ori.z_val
                            
                            # Roll (x-axis rotation)
                            sinr_cosp = 2 * (w * x + y * z)
                            cosr_cosp = 1 - 2 * (x * x + y * y)
                            drone_roll = math.degrees(math.atan2(sinr_cosp, cosr_cosp))
                            
                            # Pitch (y-axis rotation)
                            sinp = 2 * (w * y - z * x)
                            if abs(sinp) >= 1:
                                drone_pitch = math.degrees(math.copysign(math.pi / 2, sinp))
                            else:
                                drone_pitch = math.degrees(math.asin(sinp))
                            
                            # Yaw (z-axis rotation)
                            siny_cosp = 2 * (w * z + x * y)
                            cosy_cosp = 1 - 2 * (y * y + z * z)
                            drone_yaw = math.degrees(math.atan2(siny_cosp, cosy_cosp))

                        # RC Control inputs (orientation control)
                        rc_throttle = rc.throttle if rc else ''
                        # rc_roll = getattr(rc, 'roll', '') if rc else ''  # Use roll instead of steering for drone
                        # rc_pitch = getattr(rc, 'pitch', '') if rc else ''  # Use pitch instead of brake for drone
                        # rc_yaw = getattr(rc, 'yaw', '') if rc else ''  # Use yaw instead of handbrake for drone

                        # Write simplified row with all data
                        ws.append([
                            frames_written, float(f"{video_time:.3f}"), float(f"{now:.6f}"),
                            getattr(pos, 'x_val', ''), getattr(pos, 'y_val', ''), getattr(pos, 'z_val', ''),
                            getattr(vel, 'x_val', ''), getattr(vel, 'y_val', ''), getattr(vel, 'z_val', ''),
                            gimbal_roll, gimbal_pitch, gimbal_yaw,
                            drone_roll, drone_pitch, drone_yaw, rc_throttle
                        ])
                    except Exception as sensor_log_err:
                        print(f"[SensorLog] Lỗi khi ghi dữ liệu cảm biến: {sensor_log_err}")
            else:
                time.sleep(0.01)

            # --- Điều chỉnh nhịp độ và Tính toán FPS ---
            time_until_next_write_cycle = next_video_frame_write_time - time.time()
            if time_until_next_write_cycle > 0:
                time.sleep(time_until_next_write_cycle)
            
            if new_frame_successfully_acquired:
                capture_fps_frame_count += 1
            
            elapsed_time_fps_calc = time.time() - capture_fps_start_time
            if elapsed_time_fps_calc >= 1.0:
                current_capture_fps = capture_fps_frame_count / elapsed_time_fps_calc
                elapsed_recording_seconds = time.time() - start_recording_time
                print(f"FPS AirSim: {current_capture_fps:.2f} | FPS Video: {PLAYBACK_FPS:.1f} | Khung hình: {frames_written} | Thời gian: {elapsed_recording_seconds:.1f}giây ", end='\r')
                capture_fps_frame_count = 0
                capture_fps_start_time = time.time()
            
    except KeyboardInterrupt:
        print("\n\nQuá trình ghi bị người dùng ngắt (Ctrl+C).")
    finally:
        print("\r" + " " * 120 + "\r", end="")
        if VIDEO_WRITER is not None and VIDEO_WRITER.isOpened():
            print(f"Đang giải phóng trình ghi video. Tổng số khung hình đã ghi: {frames_written} vào {globals().get('OUTPUT_FILENAME', 'airsim_video.mp4')}")
            VIDEO_WRITER.release()
        # Auto-adjust column widths for better visualization
        for col in ws.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 20)  # Cap width at 20
        wb.save(sensor_log_filename)
        print(f"Đang lưu log cảm biến vào {sensor_log_filename}")
        print("Quá trình ghi đã kết thúc.")

if __name__ == '__main__':
    print("Đang cố gắng kết nối với AirSim...")
    connect_to_airsim()

    if AIRSIM_CLIENT:
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        image_type_name = get_image_type_name(IMAGE_TYPE)
        globals()['OUTPUT_FILENAME'] = f"airsim_rec_{image_type_name}_fps{int(PLAYBACK_FPS)}_{timestamp}.mp4"
        
        while True:
            command = input("Nhập 'start' để bắt đầu ghi, hoặc 'exit' để thoát: ").strip().lower()
            if command == "start":
                record_airsim_video()
                print("Quay trở lại dấu nhắc lệnh.")
            elif command == "exit":
                print("Đang thoát chương trình.")
                break
            else:
                print("Lệnh không hợp lệ.")
    else:
        print("Không thể kết nối với AirSim. Đang thoát.")